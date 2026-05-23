from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

NHS = "005EB8"; LIGHT = "DCE9F6"; GREY = "F2F2F2"; AMBER = "FFF2CC"
hdr = Font(name="Arial", bold=True, color="FFFFFF", size=11)
hfill = PatternFill("solid", fgColor=NHS)
bold = Font(name="Arial", bold=True, size=11)
base = Font(name="Arial", size=10)
ital = Font(name="Arial", italic=True, size=10, color="555555")
wrap = Alignment(wrap_text=True, vertical="top")
ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin = Side(style="thin", color="BBBBBB")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# id, title, target specialty
scenarios = [
    ("S8",  "Neonatal early-onset sepsis",        "Neonatology / NICU"),
    ("S9",  "Paediatric DKA (new T1DM)",          "Paediatrics"),
    ("S10", "Emergency LSCS + PPH",               "Obstetrics"),
    ("S11", "Polytrauma (RTC)",                   "Trauma & Orthopaedics"),
    ("S12", "Care of the Elderly (fall/delirium)","Care of the Elderly / Geriatrics"),
    ("S13", "Prolonged ITU (nec fasc)",           "Intensive Care"),
    ("S14", "Thrombolysed stroke",                "Stroke Medicine"),
    ("S15", "COPD + NIV (pre-existing DNACPR)",   "Respiratory Medicine"),
    ("S16", "Laparotomy + stoma",                 "General Surgery"),
    ("S17", "Upper GI bleed (NSAID ulcer)",       "Gastroenterology"),
    ("S18", "First unprovoked seizure",           "Neurology"),
]

wb = Workbook()

# ---------- Guide sheet ----------
g = wb.active; g.title = "Guide"
g.sheet_view.showGridLines = False
g.column_dimensions["A"].width = 3
g.column_dimensions["B"].width = 30
g.column_dimensions["C"].width = 82

g["B2"] = "AI Discharge Summary Assistant — Clinician Reviewer Scoring Sheet"
g["B2"].font = Font(name="Arial", bold=True, size=14, color=NHS)
g["B3"] = "All data is synthetic. This is YOUR central collation sheet for the clinician reviews."
g["B3"].font = ital

def grow(r, b, c):
    g[f"B{r}"] = b; g[f"B{r}"].font = bold; g[f"B{r}"].alignment = wrap
    g[f"C{r}"] = c; g[f"C{r}"].font = base; g[f"C{r}"].alignment = wrap

rows = [
 ("Method", "One clinician reviews ONE scenario — the one in their specialty — using the matching single-scenario Word pack. They mark up the grid (and the free-text box); you transcribe their ratings and comments into that scenario's row here, with their name, specialty and the date. Each row is therefore a different reviewer on a different case (not several clinicians on the same case)."),
 ("How to use", "Open the ‘Scoring’ tab. Set the Status for each row as you go (Reviewed / Awaiting reviewer / No reviewer found). Pick ratings from the dropdowns and paste in comments. The summary block auto-counts as you fill it in."),
 ("Rating scale", "Good = correct / safe.    Concern = imperfect, would change but not dangerous.    Wrong = clinically incorrect, unsafe, or fabricated."),
 ("Completeness", "Is any clinically important field missing (diagnosis, a discharge medication, follow-up, GP action, safety advice)?"),
 ("Accuracy / fabrication", "Is anything asserted that the notes do NOT support? Inventing a resus status, drug, dose, or diagnosis is the most serious failure."),
 ("Resuscitation status", "Correct, and any change during admission clearly flagged? An honest ‘Not documented’ is correct when the notes are silent — that is NOT a miss."),
 ("Medications", "Are the discharge meds and the changes (new / stopped / withheld / continued) correct and safe? Watch high-harm cases (anticoagulants, antiplatelets)."),
 ("Sign as draft?", "Bottom line: would the reviewer put their name to this as a draft for sign-off? Yes / No."),
 ("If no reviewer is found", "It is fine to leave a scenario unreviewed — mark its Status ‘No reviewer found’. Clinician review is an extra assurance layer on top of the automated cold-run + reference answer, not a gate. Options: ask a clinician in an adjacent specialty (e.g. acute/general medicine can cover several medical cases); review it yourself as the project clinician (note it as self-review in the Reviewer column); or leave it and record it as a known limitation. Partial, honest coverage beats forcing a poor-fit reviewer."),
 ("Note", "The patient-facing version's reading age is measured automatically by formula, so it is not scored here. Focus on the clinical content."),
 ("Why it matters", "This filled-in sheet is your record of INDEPENDENT clinician scoring — the rigour gap still open in EVAL_RESULTS.md §7. A recent cold review already caught five errors in our own answer key."),
]
r = 5
for b, c in rows:
    grow(r, b, c); g.row_dimensions[r].height = 14*max(1, (len(c)//95)+1); r += 1

# ---------- Scoring sheet ----------
s = wb.create_sheet("Scoring")
s.sheet_view.showGridLines = False
headers = ["Scenario", "Specialty (target)", "Status", "Completeness",
           "Accuracy /\nno fabrication", "Resuscitation\nstatus", "Medications",
           "Sign as\ndraft?", "Errors / omissions noted", "Reviewer (name)", "Date"]
widths = [30, 22, 18, 13, 14, 13, 13, 10, 40, 17, 12]
for i, (h, w) in enumerate(zip(headers, widths), start=1):
    c = s.cell(row=1, column=i, value=h); c.font = hdr; c.fill = hfill; c.alignment = ctr; c.border = border
    s.column_dimensions[chr(64+i)].width = w
s.row_dimensions[1].height = 30

# method note row under header
note = s.cell(row=2, column=1,
    value="Method: one clinician per scenario (their specialty). Set Status as you go; mark ‘No reviewer found’ if a scenario can't be placed. See the Guide tab.")
note.font = ital; note.alignment = wrap
s.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
s.cell(row=2, column=1).fill = PatternFill("solid", fgColor=AMBER)
s.row_dimensions[2].height = 26

first = 3
for idx, (sid, title, spec) in enumerate(scenarios):
    rr = first + idx
    s.cell(row=rr, column=1, value=f"{sid} — {title}").font = bold
    s.cell(row=rr, column=2, value=spec).font = base
    for col in range(1, 12):
        cell = s.cell(row=rr, column=col); cell.border = border; cell.alignment = wrap
    if idx % 2 == 0:
        for col in range(1, 12):
            s.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=GREY)
    s.row_dimensions[rr].height = 30
last = first + len(scenarios) - 1

rate_dv = DataValidation(type="list", formula1='"Good,Concern,Wrong"', allow_blank=True)
yn_dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
status_dv = DataValidation(type="list", formula1='"Reviewed,Awaiting reviewer,No reviewer found"', allow_blank=True)
s.add_data_validation(rate_dv); s.add_data_validation(yn_dv); s.add_data_validation(status_dv)
status_dv.add(f"C{first}:C{last}")
for col in ["D", "E", "F", "G"]:
    rate_dv.add(f"{col}{first}:{col}{last}")
yn_dv.add(f"H{first}:H{last}")

# ---------- Summary block ----------
sr = last + 2
s.cell(row=sr, column=1, value="SUMMARY (auto-counts as you fill in)").font = bold
dim_cols = [(4,"Completeness"),(5,"Accuracy"),(6,"Resus"),(7,"Medications")]
for col, name in dim_cols:
    L = chr(64+col)
    s.cell(row=sr, column=col, value=name).font = bold; s.cell(row=sr, column=col).alignment = ctr; s.cell(row=sr, column=col).border = border
    s.cell(row=sr+1, column=col, value=f'=COUNTIF({L}{first}:{L}{last},"Concern")').font = base
    s.cell(row=sr+2, column=col, value=f'=COUNTIF({L}{first}:{L}{last},"Wrong")').font = base
    s.cell(row=sr+3, column=col, value=f'=COUNTIF({L}{first}:{L}{last},"Good")').font = base
    for off in range(1,4):
        c = s.cell(row=sr+off, column=col); c.alignment = ctr; c.border = border
s.cell(row=sr+1, column=1, value="Concerns (⚠) flagged").font = base
s.cell(row=sr+2, column=1, value="Wrong (✗) flagged").font = base
s.cell(row=sr+3, column=1, value="Good (✓) — clean").font = base
# Sign? counts (col H = 8)
s.cell(row=sr, column=8, value="Sign?").font = bold; s.cell(row=sr, column=8).alignment = ctr; s.cell(row=sr, column=8).border = border
s.cell(row=sr+1, column=8, value=f'=COUNTIF(H{first}:H{last},"No")').font = base
s.cell(row=sr+2, column=8, value=f'=COUNTIF(H{first}:H{last},"Yes")').font = base
for off in (1,2):
    c = s.cell(row=sr+off, column=8); c.alignment = ctr; c.border = border
# Coverage (Status col C = 3)
s.cell(row=sr, column=3, value="Coverage").font = bold; s.cell(row=sr, column=3).alignment = ctr; s.cell(row=sr, column=3).border = border
s.cell(row=sr+1, column=3, value=f'=COUNTIF(C{first}:C{last},"Reviewed")&" / {len(scenarios)} reviewed"').font = base
s.cell(row=sr+2, column=3, value=f'=COUNTIF(C{first}:C{last},"No reviewer found")&" no reviewer"').font = base
for off in (1,2):
    c = s.cell(row=sr+off, column=3); c.alignment = ctr; c.border = border

s.cell(row=sr+5, column=1, value="Tip: any scenario that comes back ‘Wrong’ is worth a second reviewer before changing the reference answer (gold). Add rows below if you also review the seed or adversarial scenarios.").font = ital

s.freeze_panes = "A3"
wb.save("Reviewer-Scoring-Sheet.xlsx")
print("saved")
